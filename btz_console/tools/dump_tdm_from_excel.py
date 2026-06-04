import os

from openpyxl import load_workbook


EXCEL_PATH = r"C:\Users\Usuario\AppData\Roaming\pyRevit\Extensions\btz.extension\resources\2026-03-27 EQUIPOS ACTUALIZADOS.xlsx"


def main() -> None:
    if not os.path.exists(EXCEL_PATH):
        print("ERROR\tExcel no encontrado en ruta esperada:", EXCEL_PATH)
        return

    wb = load_workbook(EXCEL_PATH, data_only=True)

    # Usamos la primera hoja como maestro de equipos (ajustable si hace falta)
    ws = wb[wb.sheetnames[0]]

    # Detectar encabezados en la primera fila
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip().upper()] = cell.column

    print("ENCABEZADOS\t", headers)

    # Columnas candidatas para padre/nodo superior y descripción
    padre_cols = [
        c
        for name, c in headers.items()
        if any(key in name for key in ("PADRE", "PARENT", "NODO SUPERIOR", "SUPERIOR"))
    ]

    desc_cols = [
        c
        for name, c in headers.items()
        if any(key in name for key in ("DESC", "NOMBRE", "DESCRIP"))
    ]

    print("COLUMNAS_PADRE\t", padre_cols)
    print("COLUMNAS_DESC\t", desc_cols)

    # Para el código TE-TDM, buscamos en TODAS las columnas de texto de cada fila
    all_cols = list(range(1, ws.max_column + 1))

    def first_col_value(row, cols):
        for col in cols:
            if not col:
                continue
            if col - 1 >= len(row):
                continue
            val = row[col - 1]
            if isinstance(val, str):
                val = val.strip()
            if val:
                return val
        return None

    print("TE_TDM_ROWS_BEGIN")
    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo = None
        codigo_col = None
        for idx in all_cols:
            if idx - 1 >= len(row):
                continue
            val = row[idx - 1]
            if not isinstance(val, str):
                continue
            v = val.strip()
            if v.startswith("TE-TDM"):
                codigo = v
                codigo_col = idx
                break

        if not codigo:
            continue

        padre = first_col_value(row, padre_cols) or ""
        desc = first_col_value(row, desc_cols) or ""

        # TSV: CODIGO \t PADRE \t DESCRIPCION \t COL_CODIGO
        print(f"{codigo}\t{padre}\t{desc}\tCOL{codigo_col}")

    print("TE_TDM_ROWS_END")


if __name__ == "__main__":
    main()


